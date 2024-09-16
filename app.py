from typing import Optional
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def apply_discount(
    filename: str,
    sheet_name: Optional[str],
    discount_percentage: float
) -> None:
    wb = xl.load_workbook(filename)
    sheet = wb[sheet_name] if sheet_name else wb.active

    if sheet is None:
        print(f'Sheet {sheet_name} not found in file {filename}')
        return

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * discount_percentage
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    sheet.cell(1, 4, 'Discounted Price')

    values = Reference(
        sheet,
        min_row=2,
        max_row=sheet.max_row,
        min_col=4,
        max_col=4)

    chart = BarChart()
    chart.add_data(values)
    chart.title = 'Discounted Prices'
    chart.x_axis.title = 'Products'
    chart.y_axis.title = 'Prices'

    chart_cell = f'{chr(65 + sheet.max_column)}2'

    if sheet.max_row > 1:
        sheet.add_chart(chart, chart_cell)

    wb.save(filename)
    print(f'File {filename} saved with discount applied')


apply_discount(
    filename='transactions.xlsx',
    sheet_name='Planilha2',
    discount_percentage=0.9)
